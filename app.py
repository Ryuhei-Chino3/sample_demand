import streamlit as st
import pandas as pd
import numpy as np
from io import BytesIO
import re
from openpyxl import load_workbook
from openpyxl.chart import LineChart, Reference

st.set_page_config(page_title="30分値リスケーリング（全時間帯）", layout="wide")

st.title("サンプル30分値リスケーリングアプリ（全時間帯 + デマンドカーブ）")
st.markdown(
    """
    目的：サンプルの30分値データ（横に時間帯列を持つ形式）を、月ごとの合計使用量に合わせて
    **全ての時間帯列を同一比率でスケーリング**し、さらに各月の平均的なデマンドカーブをグラフ付きで
    出力する Excel ファイルを生成します。表示は小数点第一位で四捨五入しています。
    """
)

# --- ファイルアップロード ---
st.sidebar.header("1. サンプルファイルをアップロード")
uploaded_file = st.sidebar.file_uploader(
    "CSV または XLSX をアップロード", type=["csv", "xlsx"], accept_multiple_files=False
)
if uploaded_file is None:
    st.warning("まずはサンプルの30分値データ（CSV または XLSX）をアップロードしてください。")
    st.stop()

# --- 読み込み ---
@st.cache_data
def load_sample(file) -> pd.DataFrame:
    name = file.name.lower()
    if name.endswith(".csv"):
        return pd.read_csv(file)
    elif name.endswith((".xls", ".xlsx")):
        try:
            return pd.read_excel(file)
        except ImportError as ie:
            raise RuntimeError(
                "Excel 読み込みに必要なライブラリ（openpyxl）が見つかりません。"
                " requirements.txt に `openpyxl` を追加して再デプロイしてください。"
            ) from ie
    else:
        raise ValueError("対応していないファイル形式です。CSVかXLSXをアップロードしてください。")

try:
    df_raw = load_sample(uploaded_file)
except Exception as e:
    st.error(f"ファイル読み込みでエラーが発生しました: {e}")
    st.stop()

st.subheader("アップロードされた元データの先頭（構造確認）")
st.dataframe(df_raw.head())

# --- 日付/日時列指定 ---
st.sidebar.header("2. 日付列と時間帯列の指定")

possible_date_cols = []
for col in df_raw.columns:
    try:
        parsed = pd.to_datetime(df_raw[col], errors="coerce")
        if parsed.notna().mean() > 0.5:
            possible_date_cols.append(col)
    except Exception:
        continue

default_date_idx = 0
if possible_date_cols:
    try:
        default_date_idx = int(df_raw.columns.get_indexer([possible_date_cols[0]])[0])
    except Exception:
        default_date_idx = 0

date_col = st.sidebar.selectbox(
    "日付／日時列を選択してください", options=df_raw.columns.tolist(), index=default_date_idx
)

df = df_raw.copy()
try:
    df[date_col] = pd.to_datetime(df[date_col])
except Exception as e:
    st.error(f"{date_col} を日時に変換できませんでした。形式を確認してください。詳細: {e}")
    st.stop()

# --- 時間帯列の自動検出（00:00:00 形式優先、フォールバックあり） ---
expected_time_labels_hms = [f"{h:02d}:{m:02d}:00" for h in range(24) for m in (0, 30)]
candidate_time_cols = [c for c in df.columns if str(c).strip() in expected_time_labels_hms]
if not candidate_time_cols:
    pattern = re.compile(r"^([01]?\d|2[0-3]):[0-5]\d(:00)?$")
    candidate_time_cols = [c for c in df.columns if pattern.match(str(c).strip())]

st.sidebar.markdown("**自動検出された時間帯列（必要なら調整）**")
time_cols = st.sidebar.multiselect(
    "スケーリング対象とする時間帯列を選択（通常はすべて選択）",
    options=df.columns.tolist(),
    default=candidate_time_cols if candidate_time_cols else []
)

if not time_cols:
    st.error("時間帯列が選択されていません。手動で 00:00:00 ～ 23:30:00 相当の列を選択してください。")
    st.stop()

# --- 月ごとの集計 ---
df["__year_month"] = df[date_col].dt.to_period("M")
df["_row_total"] = df[time_cols].sum(axis=1)

monthly_original = (
    df.groupby("__year_month")["_row_total"]
    .sum()
    .rename("元の月合計")
    .to_frame()
)
monthly_original["表示用月"] = monthly_original.index.to_timestamp()

# --- 月使用量入力（年月ラベル1行、その下に入力欄） ---
st.subheader("各月の元の合計使用量と新しい月合計の入力")
st.markdown("年月ラベルの下に目標月使用量を横並びで入力してください。元の月合計はラベル下に小さく表示（小数点第一位）。")

with st.form("monthly_targets_form"):
    periods = list(monthly_original.index)
    labels = [p.strftime("%Y/%-m") if hasattr(p, "strftime") else str(p) for p in periods]
    cols = st.columns(len(periods))
    target_inputs = {}
    for col, period, label in zip(cols, periods, labels):
        with col:
            st.markdown(f"**{label}**")
            orig = monthly_original.loc[period, "元の月合計"]
            st.markdown(f"<div style='font-size:0.75rem; color:gray;'>元の合計: {orig:.1f}</div>", unsafe_allow_html=True)
            user_str = st.text_input(
                label="",
                value="",
                placeholder="例: 1234.5",
                key=f"target_{period}",
                help="目標の月合計使用量（空白なら元の値）",
            )
            if user_str.strip() == "":
                target_inputs[period.strftime("%Y-%m")] = float(orig)
            else:
                try:
                    target_inputs[period.strftime("%Y-%m")] = float(user_str.replace(",", ""))
                except ValueError:
                    st.warning(f"{label} の入力が数値に解釈できません。元の合計を使います。")
                    target_inputs[period.strftime("%Y-%m")] = float(orig)
    submitted = st.form_submit_button("スケーリング実行")

if not submitted:
    st.info("上の行に各月の目標使用量を入力して「スケーリング実行」を押してください。")
    st.stop()

# --- スケーリング ---
scaling = {}
for label, target in target_inputs.items():
    period = pd.Period(label, freq="M")
    orig = monthly_original.loc[period, "元の月合計"]
    if orig == 0:
        scaling[period] = np.nan
    else:
        scaling[period] = target / orig

df_scaled = df.copy()
df_scaled["__scale_factor"] = df_scaled["__year_month"].map(scaling)

for col in time_cols:
    df_scaled[col] = df_scaled[col] * df_scaled["__scale_factor"].where(df_scaled["__scale_factor"].notna(), 1.0)

# --- デマンドカーブ（各月の平均プロファイル）作成 ---
monthly_profile = (
    df_scaled.groupby("__year_month")[time_cols]
    .mean()
)
profile_df = monthly_profile.T
profile_df.columns = [p.strftime("%Y-%m") for p in profile_df.columns]

# --- 検証表示 ---
st.subheader("スケーリング後の各月合計の検証")
scaled_monthly = (
    df_scaled.groupby("__year_month")[time_cols]
    .sum()
    .sum(axis=1)
    .rename("スケーリング後合計")
    .to_frame()
)
compare = monthly_original.join(scaled_monthly)
compare["入力目標"] = [target_inputs[p.strftime("%Y-%m")] for p in compare.index]
compare["比率（実績/目標）"] = compare["スケーリング後合計"] / compare["入力目標"].replace({0: np.nan})
st.dataframe(compare.style.format({
    "元の月合計": "{:.1f}",
    "スケーリング後合計": "{:.1f}",
    "入力目標": "{:.1f}",
    "比率（実績/目標）": "{:.4f}"
}))

for period, row in compare.iterrows():
    if row["入力目標"] == 0 and row["元の月合計"] != 0:
        st.warning(f"{period.strftime('%Y-%m')} の入力目標が0です。元の値があるため、全て0になります。")
    if row["元の月合計"] == 0:
        st.warning(f"{period.strftime('%Y-%m')} は元の合計が0のためスケーリングされていません。")

# --- 出力 ---
st.subheader("出力ファイル設定とダウンロード")
output_name = st.text_input("出力ファイル名（.xlsx で終わる）", value="rescaled_30min_full_with_curve.xlsx")
if not output_name.lower().endswith(".xlsx"):
    st.error("ファイル名は .xlsx で終わる必要があります。")
    st.stop()

def to_excel_bytes_with_curve(df_rescaled: pd.DataFrame, curve_df: pd.DataFrame):
    output = BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        to_export_main = df_rescaled.drop(columns=["__year_month", "_row_total", "__scale_factor"], errors="ignore")
        to_export_main.to_excel(writer, index=False, sheet_name="Rescaled")
        curve_df_reset = curve_df.reset_index().rename(columns={"index": "時間帯"})
        curve_df_reset.to_excel(writer, index=False, sheet_name="DemandCurve")
    output.seek(0)
    wb = load_workbook(filename=output)
    ws = wb["DemandCurve"]

    max_row = ws.max_row
    max_col = ws.max_column

    chart = LineChart()
    chart.title = "月ごとの平均デマンドカーブ（時間帯ごと）"
    chart.y_axis.title = "平均使用量"
    chart.x_axis.title = "時間帯"

    cats = Reference(ws, min_col=1, min_row=2, max_row=max_row)
    for col_idx in range(2, max_col + 1):
        data = Reference(ws, min_col=col_idx, min_row=2, max_row=max_row)
        chart.add_data(data, titles_from_data=False)
        header_value = ws.cell(row=1, column=col_idx).value
        if chart.series:
            chart.series[-1].title = header_value

    chart.set_categories(cats)
    ws.add_chart(chart, f"B{max_row + 2}")

    out_bytes = BytesIO()
    wb.save(out_bytes)
    return out_bytes.getvalue()

try:
    excel_bytes = to_excel_bytes_with_curve(df_scaled, profile_df)
except ImportError:
    st.error("Excel 出力に必要なライブラリが見つかりません。openpyxl を requirements.txt に追加してください。")
    st.stop()

st.download_button(
    label="スケーリング結果（グラフ付き）を Excel ダウンロード",
    data=excel_bytes,
    file_name=output_name,
    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
)

st.success("処理完了。出力ファイルをダウンロードしてください。")
